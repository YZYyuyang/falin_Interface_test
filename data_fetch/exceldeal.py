import openpyxl,yaml
from openpyxl.styles import colors
from openpyxl.styles import Font
from openpyxl.styles.colors import *

class ExcelDeal(object):
    """
    处理Excel的类，读写单元格
    """
    def __init__(self):
        stream = open("..\data\yaml.yaml", mode='r', encoding="utf-8")
        data = yaml.load(stream)["path"]["excelpath"]
        print(data)
        self.datafile = data
        self.excel = openpyxl.load_workbook(self.datafile)  # 打开指定excel：（login.xlsx）
        # self.excel = openpyxl.load_workbook(datafile)  # 打开指定excel：（login.xlsx）

    # 根据sheet页名返回sheet页
    def get_sheet_by_name(self,sheetname):      #打开第一个sheet页one
        return self.excel.get_sheet_by_name(sheetname)

    # 获取sheetname页所有的行对象
    def get_rows_by_nameb(self,sheetname):
        sheet = self.excel.get_sheet_by_name(sheetname)
        # 遍历所有的行所有的列
        for row in sheet.iter_rows():   #所有的行
            for cell in row:
                pass
            #     print(f'cell.coordinate={cell.coordinate},cell.value={cell.value}')
            print(f'row={row}')

    #获取行数
    def get_rows(self,sheetname):
        sheet = self.excel.get_sheet_by_name(sheetname)
        # 遍历所有的行所有的列
        rows = sheet.max_row  # 所有的行
        print(rows)
        return rows

    #获取列数
    def get_cols(self,sheetname):
        sheet = self.excel.get_sheet_by_name(sheetname)
        # 遍历所有的行所有的列
        cols = sheet.max_column  # 所有的行
        print(cols)
        return cols

    # 获取sheetname页所有的行对象
    def get_rows_by_name(self, sheetname):
        sheet = self.excel.get_sheet_by_name(sheetname)
        rows = []
        # 遍历所有的行所有的列
        for row in sheet.iter_rows():
            rowdata = []
            for cell in row:
                rowdata.append(cell.value)
            rows.append(rowdata)
        # print(f'rows={rows}')
        # print(f'rows[0]={rows[0]}')
        # print(f'dada,rowdata={rowdata}')
        return rows

    # 获取sheetname页指定的行对象
    def get_rows_by_name_x(self, sheetname,x):
        sheet = self.excel.get_sheet_by_name(sheetname)
        rows = []
        # 遍历所有的行所有的列
        for row in sheet.iter_rows():
            rowdata = []
            for cell in row:
                rowdata.append(cell.value)
            rows.append(rowdata)
        rows_x = rows[x]
        print(f'rows[0]={rows[x]}')
        # print(f'dada,rowdata={rowdata}')
        return rows_x

    # 获取sheetname页中所有的列对象
    def get_cols_by_name(self,sheetname):
        sheet = self.excel.get_sheet_by_name(sheetname)
        colums = []
        # 遍历所有的行所有的列
        for col in sheet.iter_cols():
            columsdata = []
            for cell in col:
                columsdata.append(cell.value)
            colums.append(columsdata)
        # print(f'colums={colums}')
        # print(f'colums[0]={colums[4]}')

    # 获取sheetname页中指定的列对象
    def get_cols_by_name_y(self, sheetname,y):
        sheet = self.excel.get_sheet_by_name(sheetname)
        colums = []
        # 遍历所有的行所有的列
        for col in sheet.iter_cols():
            columsdata = []
            for cell in col:
                columsdata.append(cell.value)
            colums.append(columsdata)
        colums_y = colums[y]
        print(f'colums[0]={colums[y]}')
        return colums_y

    # 针对指定sheet页按行列下标获取单元格内容
    def get_cell_by_row_and_col(self,sheetname,row,col):
        sheet = self.excel.get_sheet_by_name(sheetname)
        celldata = sheet.cell(row,col)
        print(f'celldata.value={celldata.value}')
        return celldata.value

    # 修改指定单元格内容
    def write_cell_by_row_and_col(self,sheetname,row,col,data): #成功的用例
        sheet = self.excel.get_sheet_by_name(sheetname)
        celldata = sheet.cell(row,col)
        # 设置字体与颜色
        celldata.font = Font(color=colors.GREEN,bold=True)
        celldata.value = data
        # 保存修改
        self.excel.save(self.datafile)
        print(f'修改成功')

    def write_cell_by_row_and_col_False(self,sheetname,row,col,data):#失败的用例
        sheet = self.excel.get_sheet_by_name(sheetname)
        celldata = sheet.cell(row,col)
        # 设置字体与颜色
        celldata.font = Font(color=colors.RED,bold=True)
        celldata.value = data
        # 保存修改
        self.excel.save(self.datafile)
        print(f'修改成功')

if __name__ == '__main__':
    excel = ExcelDeal()
    # excel.write_cell_by_row_and_col(sheetname = "Sheet1",row = 1,col = 1,data = '描述1')
    # excel.get_rows_by_name_x(sheetname = "Sheet1",x = 0)
    # excel.get_cols_by_name_y(sheetname = "Sheet1",y=0)
    excel.get_cols(sheetname = "Sheet1")

